"""File parsers for different file types."""

from __future__ import annotations

import io
from abc import ABC, abstractmethod
from pathlib import Path


class FileParser(ABC):
    """Base class for file parsers."""

    @abstractmethod
    def parse(self, content_bytes: bytes) -> str:
        """
        Parse file content and return text.

        Args:
            content_bytes: Raw file content as bytes

        Returns:
            Extracted text content
        """

    @staticmethod
    def get_parser(file_key: str) -> FileParser:
        """
        Get appropriate parser based on file extension.

        Args:
            file_key: S3 object key (file path)

        Returns:
            Appropriate FileParser instance
        """
        extension = Path(file_key).suffix.lower()

        if extension == ".pdf":
            return PDFFileParser()
        elif extension in [".xlsx", ".xls"]:
            return ExcelFileParser()
        elif extension == ".csv":
            return CSVFileParser()
        else:  # .txt, .md, etc.
            return TextFileParser()


class TextFileParser(FileParser):
    """Parser for text-based files."""

    def parse(self, content_bytes: bytes) -> str:
        """
        Parse text file with multiple encoding attempts.

        Args:
            content_bytes: Raw file content

        Returns:
            Decoded text content
        """
        try:
            return content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return content_bytes.decode("shift_jis")
            except UnicodeDecodeError:
                return content_bytes.decode("latin-1")


class PDFFileParser(FileParser):
    """Parser for PDF files."""

    def parse(self, content_bytes: bytes) -> str:
        """
        Extract text from PDF file.

        Args:
            content_bytes: Raw PDF file content

        Returns:
            Extracted text from all pages

        Raises:
            Exception: If PDF parsing fails
        """
        try:
            import PyPDF2

            pdf_file = io.BytesIO(content_bytes)
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            text_parts = []
            for page_num, page in enumerate(pdf_reader.pages, 1):
                page_text = page.extract_text()
                if page_text.strip():
                    text_parts.append(f"=== Page {page_num} ===\n{page_text}")

            if not text_parts:
                return "[No extractable text found in PDF]"

            return "\n\n".join(text_parts)

        except Exception as e:
            raise Exception(f"Failed to parse PDF file: {str(e)}") from e


class ExcelFileParser(FileParser):
    """Parser for Excel files."""

    def parse(self, content_bytes: bytes) -> str:
        """
        Extract data from Excel file.

        Args:
            content_bytes: Raw Excel file content

        Returns:
            Formatted text representation of Excel data

        Raises:
            Exception: If Excel parsing fails
        """
        try:
            import openpyxl

            excel_file = io.BytesIO(content_bytes)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)

            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")

                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                header_text = ", ".join(str(h) for h in headers if h is not None)
                text_parts.append(f"Headers: {header_text}")

                # Get sample rows (up to 10 rows for context)
                text_parts.append("\nSample rows:")
                max_sample_rows = min(10, sheet.max_row)
                for row_num in range(2, max_sample_rows + 1):
                    row_values = [cell.value for cell in sheet[row_num]]
                    row_text = ", ".join(str(v) if v is not None else "" for v in row_values)
                    text_parts.append(f"Row {row_num}: {row_text}")

                text_parts.append(f"\nTotal rows: {sheet.max_row}")
                text_parts.append(f"Total columns: {sheet.max_column}")

            return "\n".join(text_parts)

        except Exception as e:
            raise Exception(f"Failed to parse Excel file: {str(e)}") from e


class CSVFileParser(FileParser):
    """Parser for CSV files."""

    def parse(self, content_bytes: bytes) -> str:
        """
        Parse CSV file with structure information.

        Args:
            content_bytes: Raw CSV file content

        Returns:
            Structured text with headers and sample rows

        Raises:
            Exception: If CSV parsing fails
        """
        try:
            # First decode as text
            text = TextFileParser().parse(content_bytes)
            lines = [line for line in text.split("\n") if line.strip()]

            if not lines:
                return "[Empty CSV file]"

            result = []
            result.append(f"CSV Headers: {lines[0]}")
            result.append("\nSample rows (first 5):")

            # Add up to 5 sample rows
            for i, line in enumerate(lines[1:6], 1):
                result.append(f"Row {i}: {line}")

            result.append(f"\nTotal rows: {len(lines) - 1}")

            return "\n".join(result)

        except Exception as e:
            raise Exception(f"Failed to parse CSV file: {str(e)}") from e
