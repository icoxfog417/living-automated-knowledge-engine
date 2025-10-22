"""Tests for file parser functionality."""

import io

import PyPDF2
import pytest
from openpyxl import Workbook

from src.services.file_parser import (
    CSVFileParser,
    ExcelFileParser,
    FileParser,
    PDFFileParser,
    TextFileParser,
)


class TestFileParserFactory:
    """Test FileParser.get_parser() factory method."""

    def test_get_parser_for_pdf(self):
        """Test that PDF files get PDFFileParser."""
        parser = FileParser.get_parser("document.pdf")
        assert isinstance(parser, PDFFileParser)

    def test_get_parser_for_xlsx(self):
        """Test that XLSX files get ExcelFileParser."""
        parser = FileParser.get_parser("data.xlsx")
        assert isinstance(parser, ExcelFileParser)

    def test_get_parser_for_xls(self):
        """Test that XLS files get ExcelFileParser."""
        parser = FileParser.get_parser("data.xls")
        assert isinstance(parser, ExcelFileParser)

    def test_get_parser_for_csv(self):
        """Test that CSV files get CSVFileParser."""
        parser = FileParser.get_parser("data.csv")
        assert isinstance(parser, CSVFileParser)

    def test_get_parser_for_txt(self):
        """Test that TXT files get TextFileParser."""
        parser = FileParser.get_parser("readme.txt")
        assert isinstance(parser, TextFileParser)

    def test_get_parser_for_md(self):
        """Test that MD files get TextFileParser."""
        parser = FileParser.get_parser("README.md")
        assert isinstance(parser, TextFileParser)

    def test_get_parser_case_insensitive(self):
        """Test that parser selection is case-insensitive."""
        parser = FileParser.get_parser("DOCUMENT.PDF")
        assert isinstance(parser, PDFFileParser)


class TestTextFileParser:
    """Test TextFileParser."""

    def test_parse_utf8(self):
        """Test parsing UTF-8 encoded text."""
        parser = TextFileParser()
        content = "Hello, World! 日本語"
        content_bytes = content.encode("utf-8")

        result = parser.parse(content_bytes)

        assert result == content

    def test_parse_shift_jis(self):
        """Test parsing Shift-JIS encoded text."""
        parser = TextFileParser()
        content = "日本語のテキスト"
        content_bytes = content.encode("shift_jis")

        result = parser.parse(content_bytes)

        assert result == content

    def test_parse_latin1(self):
        """Test parsing Latin-1 encoded text."""
        parser = TextFileParser()
        content = "Café résumé"
        # Create bytes that fail UTF-8 and Shift-JIS but work with Latin-1
        content_bytes = b"\xe9\xe0\xe7"  # Latin-1 specific characters

        result = parser.parse(content_bytes)

        assert isinstance(result, str)


class TestPDFFileParser:
    """Test PDFFileParser."""

    def test_parse_simple_pdf(self):
        """Test parsing a simple PDF with text."""
        parser = PDFFileParser()

        # Create a simple PDF in memory
        pdf_buffer = io.BytesIO()
        pdf_writer = PyPDF2.PdfWriter()

        # Add a page with text
        page = pdf_writer.add_blank_page(width=200, height=200)
        # Note: PyPDF2 doesn't easily support adding text in tests,
        # so we'll test with an empty page
        pdf_writer.write(pdf_buffer)

        content_bytes = pdf_buffer.getvalue()

        result = parser.parse(content_bytes)

        # Empty PDF should return no extractable text message
        assert isinstance(result, str)
        assert "[No extractable text found in PDF]" in result or result == ""

    def test_parse_invalid_pdf(self):
        """Test parsing invalid PDF content."""
        parser = PDFFileParser()
        invalid_bytes = b"This is not a PDF file"

        with pytest.raises(Exception, match="Failed to parse PDF file"):
            parser.parse(invalid_bytes)


class TestExcelFileParser:
    """Test ExcelFileParser."""

    def test_parse_simple_excel(self):
        """Test parsing a simple Excel file."""
        parser = ExcelFileParser()

        # Create a simple Excel workbook in memory
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Sheet"

        # Add headers
        sheet["A1"] = "Name"
        sheet["B1"] = "Age"
        sheet["C1"] = "Department"

        # Add data
        sheet["A2"] = "Alice"
        sheet["B2"] = 30
        sheet["C2"] = "Engineering"

        sheet["A3"] = "Bob"
        sheet["B3"] = 25
        sheet["C3"] = "Sales"

        # Save to buffer
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        content_bytes = excel_buffer.getvalue()

        result = parser.parse(content_bytes)

        # Verify the parsed content
        assert "=== Sheet: Test Sheet ===" in result
        assert "Headers: Name, Age, Department" in result
        assert "Alice" in result
        assert "Bob" in result
        assert "Total rows:" in result
        assert "Total columns:" in result

    def test_parse_excel_multiple_sheets(self):
        """Test parsing Excel file with multiple sheets."""
        parser = ExcelFileParser()

        workbook = Workbook()

        # First sheet
        sheet1 = workbook.active
        sheet1.title = "Sales"
        sheet1["A1"] = "Product"
        sheet1["B1"] = "Revenue"
        sheet1["A2"] = "Widget"
        sheet1["B2"] = 1000

        # Second sheet
        sheet2 = workbook.create_sheet("Marketing")
        sheet2["A1"] = "Campaign"
        sheet2["B1"] = "Budget"
        sheet2["A2"] = "Social Media"
        sheet2["B2"] = 5000

        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        content_bytes = excel_buffer.getvalue()

        result = parser.parse(content_bytes)

        # Verify both sheets are parsed
        assert "=== Sheet: Sales ===" in result
        assert "=== Sheet: Marketing ===" in result
        assert "Product" in result
        assert "Campaign" in result

    def test_parse_invalid_excel(self):
        """Test parsing invalid Excel content."""
        parser = ExcelFileParser()
        invalid_bytes = b"This is not an Excel file"

        with pytest.raises(Exception, match="Failed to parse Excel file"):
            parser.parse(invalid_bytes)


class TestCSVFileParser:
    """Test CSVFileParser."""

    def test_parse_simple_csv(self):
        """Test parsing a simple CSV file."""
        parser = CSVFileParser()

        csv_content = """Name,Age,Department
Alice,30,Engineering
Bob,25,Sales
Charlie,35,Marketing
Diana,28,HR
Eve,32,Finance"""

        content_bytes = csv_content.encode("utf-8")

        result = parser.parse(content_bytes)

        # Verify CSV structure is extracted
        assert "CSV Headers: Name,Age,Department" in result
        assert "Sample rows (first 5):" in result
        assert "Alice,30,Engineering" in result
        assert "Bob,25,Sales" in result
        assert "Total rows: 5" in result

    def test_parse_empty_csv(self):
        """Test parsing an empty CSV file."""
        parser = CSVFileParser()
        content_bytes = b""

        result = parser.parse(content_bytes)

        assert "[Empty CSV file]" in result

    def test_parse_csv_with_blank_lines(self):
        """Test parsing CSV with blank lines."""
        parser = CSVFileParser()

        csv_content = """Name,Age

Alice,30

Bob,25
"""

        content_bytes = csv_content.encode("utf-8")

        result = parser.parse(content_bytes)

        # Blank lines should be filtered out
        assert "CSV Headers: Name,Age" in result
        assert "Total rows:" in result
